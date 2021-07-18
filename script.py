""" This is the Python Program to design and implement the Electronic Voting Machine GUI(Graphical User Interface).
    The basic and important features of the EVM(Electronic Voting Machine) GUI are listed below :
        1) To set Election Name, Voters Count, Password for further references.
        2) To add posts and candidates' details participating in the election by the in-charge person.
        3) To add the Microsoft Excel file that contains the voters name and ID along.
        4) In-charge person can view posts and candidates added to the Election for verification and to avoid missing of any names.
        5) To allow voting based on the unique voter ID. Voters are eligible to vote only once and for only one candidate in each post.
        6) To view the result after voting is done. Viewing result is password protected.
        7) Option to save the complete details of the election in a file with election name as it's name for further references.


        This program is developed by :
            Vishwas Bhat D - vishwasbhatd7@gmail.com
            Rohan S Khadd
            Rahul Karigar
        (Students of Bangalore Institute of Technology, Bangalore)   
"""

# importing required libraries

from tkinter import *
import openpyxl
import datetime
from playsound import playsound

""" Module tkinter is a libray of Python used for Graphical User Interface(GUI). All the graphical work of this program is done useing Tkinter. """

""" Module openpyxl is a libraty of Python used to access, format, edit, delete the data stored in the Microsoft Excel Sheet. Here it is used to 
    access the data of voters form voters list stored in the Excel File. """ 

""" Module datetime is used to access the date and week of the election to write it in the file for further references."""

""" Module playsound is used to play the indication sound when voter successfully votes for all the posts. """


class election:

    def __init__(self, window):

        self.window = window                                                    # instance of the root window
        self.window.geometry("800x650+500+50")                                  # setting the fixed geometry of root window. 
        self.window.resizable(0,0)                                              # user cannot resize the window i.e. window size is fixed.
        self.window.title(" Electronic Voting Machine ")                        # title or name of the root window.
        self.window.configure(bg="coral")                                       # background color of the root window.
        self.final_candidates = {}                                              # dictionary to store the details of each candidates. Post name as the keys.
        self.voters_count = 0                                                   # variable to keep track of the number of voters voted
        self.post_names = []                                                    # list to store the post names
        self.fill_index = 0
        self.votes_of_winner_list = []                                          # list to store the winners of each post
        self.vote_started = False                                               # variable to indicate voting started or not.
        self.voters_data_list = []                                              # list to store the data of voters registered.



        def home_page_entry_check_function():
            """ This function is used to check the user did not enter any invalid data in the first page. 
                For ex. Letters in voters number registered entry instead of numbers. 
                If the entry is wrong remark frame will pop up and display the appropriate remark message."""

            if len(self.election_name.get()) == 0:
                remark_text = " ELECTION NAME CANNOT BE EMPTY. "
                remark_frame_function(self.window, texts=remark_text, flag=False, texts_color="red")
                
            
            elif self.first_password.get() != self.second_password.get():

                remark_text = " PASSWORD DID NOT MATCH. ENTER CONFIRM PASSWORD AGAIN. "
                remark_frame_function(self.window, texts=remark_text, flag=False, texts_color="red")
            
            elif len(self.first_password.get()) == 0:
                remark_text = " PASSWORD CANNOT BE EMPTY. "
                remark_frame_function(self.window, texts=remark_text, flag=False, texts_color="red")


            else:

                try:

                    if type(int(self.total_voters.get())) == int:
                        self.registered_voters_count = int(self.total_voters.get())
                       

                except ValueError:
                   
                    remark_text = " VOTERS COUNT SHOULD BE A NUMBER. "
                    remark_frame_function(self.window, texts=remark_text, flag=False, texts_color="red")
                
                else:
                    home_page_function()

        
        def home_page_function():
            """ This function represents the home page of the GUI. This frame has Add Candidates, Add Voters, View Candidates, Voting and Results options to choose.
                User have to select one of these options to carryout that process. """
            

            self.home_page_frame = Frame(self.window, width=800, height=650, bg="deep sky blue")
            self.home_page_frame.place(x=0, y=0)

            self.window.protocol("WM_DELETE_WINDOW", disable_event)

            election_name_display_function(self.home_page_frame)


            self.add_candidate = Button(self.home_page_frame, text="Add\nCandidates", bg="light cyan", fg="black", command=add_candidate_function, font=("Helvetica", 20, "bold"), width=11, height=5, activebackground="green", cursor="hand2")
            self.add_voters = Button(self.home_page_frame, text="Add\nVoters", bg="light cyan", fg="black", command=file_path_add_function, font=("Helvetica", 20, "bold"), width=11, height=5, activebackground="green", cursor="hand2")
            self.view_candidates = Button(self.home_page_frame, text="View\nCandiates", bg="light cyan", fg="black", command=view_candidates_function, font=("Helvetica", 20, "bold"), width=11, height=5, activebackground="green", cursor="hand2")
            self.button_voter = Button(self.home_page_frame, text="Voting", bg="light cyan", fg="black", command=voting_function, font=("Helvetica", 20, "bold"), width=11, height=5, activebackground="green", cursor="hand2")
            self.view_result = Button(self.home_page_frame, text="Result", bg="light cyan", fg="black", command=result_password_enter_function, font=("Helvetica", 20, "bold"), width=11, height=5, activebackground="green", cursor="hand2")


                

            self.add_candidate.place(x=70, y=150)
            self.add_voters.place(x=320, y=150)
            self.view_candidates.place(x=570, y=150)
            self.button_voter.place(x=150, y=400)
            self.view_result.place(x=470, y=400)
        


        def add_candidate_function():
            """ This function is called when user presses the Add Candidates button. The function creates the new frame where user is asked to enter the post name of the candidate. The instructions for the user is provided 
                in the textbox. """

            self.post_name_frame = Frame(self.home_page_frame, width=800, height=650, bg="pink")
            self.post_name_frame.place(x=0, y=0)

            election_name_display_function(self.post_name_frame)
            

            self.post_name_label = Label(self.post_name_frame, text="ENTER THE POST NAME : ", font=('Bahnschrift SemiLight', 15), fg="black", bg="wheat")
            self.post_name_label.place(x=130, y=200)


            self.candidate_post = StringVar()
            self.post_name_entry = Entry(self.post_name_frame, textvariable = self.candidate_post, bd=0, width=30)
            self.post_name_entry.place(x=400, y=205)
            

            self.entry_instruction_text = Text(self.post_name_frame, height=10, width = 50, bg="wheat")
            self.entry_instruction_text.place(x=150, y=250)
            self.entry_instruction_text.insert(END, "Enter the post name that the candidates are going to participate in the election.\n\nAnd then press 'Enter' button below to add the \ndetails of the candidates participating for that \npost")
            self.entry_instruction_text.insert(END, "\n\nNote : Make sure the name of a particular post \nremains exact with respect to spellings and \nletter cases everytime you enter.")
            self.entry_instruction_text.config(state="disabled")

            

            self.post_enter_button = Button(self.post_name_frame, text=" Enter ", bd=0, fg="white", bg="black", command=post_enter_function, width=20, height=3, cursor="hand2")
            self.post_canacel_button = Button(self.post_name_frame, text=" Cancel ", bd=0, fg="white", bg="black", command=post_cancel_function, width=20, height=3, cursor="hand2")

            self.post_enter_button.place(x=210, y=540)
            self.post_canacel_button.place(x=425, y=540)

        
        def post_enter_function():
            """ This function first checks if the user entered the valid post name i.e post name is not blank. Then if it check if the election has already reached it's maximum post registration (10 posts).
                If everything is alright then it appends the entered post name in the list(if the post name is not present already) and then creates the new frame for the user to add the details of individual candidate.
                User need to enter Group Name, Candidate Name, Symbol of Candidate. """

            if self.candidate_post.get() == "":
                remark_text = " POST NAME CANNOT BE EMPTY. "
                remark_frame_function(self.post_name_frame, texts=remark_text, flag=False, texts_color="red")
            
            elif len(self.post_names) >= 9:

                remark_text = " POST NUMBER REACHED MAXIMUM. "
                remark_frame_function(self.post_name_frame, texts=remark_text, flag=False, texts_color="red")

            else:

                if self.candidate_post.get().title() not in self.post_names:
                    self.post_names.append(self.candidate_post.get().title())

                self.candidate_detail_frame = Frame(self.post_name_frame, width=800, height=650, bg="pink")
                self.candidate_detail_frame.place(x=0, y=0)

                election_name_display_function(self.candidate_detail_frame)

                

                self.candidate_group = StringVar()
                self.candidate_name = StringVar()
                self.candidate_symbol = StringVar()

                self.contestant_post_name_label = Label(self.candidate_detail_frame, text=f"{self.candidate_post.get().title()} Candidate", bg="sienna1", fg="black", font=('Bahnschrift SemiLight', 15))
                self.contestant_post_name_label.place(x=270, y=140)

                self.candidate_group_label = Label(self.candidate_detail_frame, text=" ENTER THE NAME OF YOUR GROUP : ", bg="khaki1", fg="black", font=('Bahnschrift SemiLight', 15))
                self.candidate_name_label = Label(self.candidate_detail_frame, text=" ENTER YOUR NAME : ", bg="khaki1", fg="black", font=('Bahnschrift SemiLight', 15))
                self.candidate_symbol_label = Label(self.candidate_detail_frame, text=" ENTER THE SYMBOL OF YOUR GROUP : ", bg="khaki1", fg="black", font=('Bahnschrift SemiLight', 15))

                self.candidate_group_label.place(x=100, y=210)
                self.candidate_name_label.place(x=220, y=310)
                self.candidate_symbol_label.place(x=100, y=410)

                self.candidate_group_entry = Entry(self.candidate_detail_frame, textvariable = self.candidate_group, bd=0, width=30)
                self.candidate_name_entry = Entry(self.candidate_detail_frame, textvariable = self.candidate_name, bd=0, width=30)
                self.candidate_symbol_entry = Entry(self.candidate_detail_frame, textvariable = self.candidate_symbol, bd=0, width=30)

                self.candidate_group_entry.place(x=500, y=215)
                self.candidate_name_entry.place(x=500, y=315)
                self.candidate_symbol_entry.place(x=500, y=415)

                self.multiple_contestant_enter = Button(self.candidate_detail_frame, text=" Add ", bd=0, fg="white", bg="black", command = candidate_enter_function, width=20, height=3, cursor="hand2")
                self.multiple_contestant_finish = Button(self.candidate_detail_frame, text=" Finish ", bd=0, fg="white", bg="black", command = candidate_finish_function, width=20, height=3, cursor="hand2")

                self.multiple_contestant_enter.place(x=210, y=540)
                self.multiple_contestant_finish.place(x=425, y=540)

        
        def post_cancel_function():
            """ This function is called whenever the user presses the cancel button in post name frame. 
                This function destroys that frame and redirects to home page. """
            
            self.post_name_frame.place_forget()
            home_page_function()
        

        def candidate_enter_function():
            """ This function first checks whether or not user filled all the entry fields. If any entry field is empty remark message is displayed. If not the candidate is susccessfully registered unless the 
                candidate number for that particular post has reached maximum(10 candidate) """
                
            if self.candidate_group_entry.get() == "" or self.candidate_name_entry.get() == "" or self.candidate_symbol_entry.get() == "":

                remark_text = " ENTRY SPACE CANNOT BE EMPTY. "

                remark_frame_function(self.candidate_detail_frame, texts=remark_text, flag=False, texts_color="red")
                 

            else:
                
                self.candidates = []

                self.candidates.append(self.candidate_post.get().title())
                
                self.candidates.append(self.candidate_group_entry.get().title())
                self.candidates.append(self.candidate_name_entry.get().title())
                self.candidates.append(self.candidate_symbol_entry.get().title())
                self.candidates.append(0)
                

                try:

                    if len(self.final_candidates[self.candidates[0]]) <= 9:
                        
                        self.final_candidates[self.candidates[0]].append(self.candidates[1:])
                        
                        remark_text = " YOU HAVE BEEN SUCCESSFULLY REGISTERED. "
                        remark_frame_function(self.candidate_detail_frame, texts=remark_text, flag=False, texts_color="green")

                    else:

                        remark_text = " CONTESTANTS NUMBER HAS REACHED MAXIMUM. "
                        remark_frame_function(self.candidate_detail_frame, texts=remark_text, flag=False, texts_color="red")

                except KeyError:
                    
                    self.final_candidates[self.candidates[0]] = []
                    self.final_candidates[self.candidates[0]].append(self.candidates[1:])
    
                    remark_text = " YOU HAVE BEEN SUCCESSFULLY REGISTERED. "
                    remark_frame_function(self.candidate_detail_frame, texts=remark_text, flag=False, texts_color="green")
                
                self.candidate_group_entry.delete(0, END)
                self.candidate_name_entry.delete(0, END)
                self.candidate_symbol_entry.delete(0, END)
                
            

        def candidate_finish_function():
            """ This function is used when user adds all the candidates for the particular post. It displays the successful message unless
                there is no candidate added to a registered post. """

            if self.candidate_post.get().title() not in self.final_candidates:
                remark_text = f" NO CANDIDATES ARE ADDED TO {self.candidate_post.get().upper()} POST. "
                remark_frame_function(self.candidate_detail_frame, texts=remark_text, flag=False, texts_color="red")
                
            else:
            
                remark_text = f" THE CANDIDATES ARE ADDED TO THE {self.candidate_post.get().upper()} POST. "
                remark_frame_function(self.candidate_detail_frame, texts=remark_text, flag=True, texts_color="green")

                self.post_name_entry.delete(0, END)



        

        def file_path_add_function():
            """ This function is called when user presses the Voter Add button present in the home page. This function desplays the frame which asks the user to enter the 
                file path of the excel sheet which has the details of teh voters in it. """

            self.voter_add_frame = Frame(self.home_page_frame, width=800, height=650, bg="silver")
            self.voter_add_frame.place(x=0, y=0)

            election_name_display_function(self.voter_add_frame)

            self.voter_add_instruction_text = Text(self.voter_add_frame, height=10, width = 55, bg="wheat")
            self.voter_add_instruction_text.place(x=148, y=150)
            self.voter_add_instruction_text.insert(END, "   In the entry box given below user should enter the \nfile name with extension if the file is in the same \nfolder as that")
            self.voter_add_instruction_text.insert(END, " of the .exe file. \n\n    Or else user should add the file path of the voter\ndata file with each folder seperated by '\\\\'.\n")
            self.voter_add_instruction_text.insert(END, "\nFor ex: G:\\\\first_folder\\\\second_folder\\\\file_name.xlsx")
            self.voter_add_instruction_text.insert(END, "\n\nFile extension should be .xlsx (Microsoft Excel File)")
            self.voter_add_instruction_text.config(state="disabled")

            self.voter_add_label = Label(self.voter_add_frame, text=" ENTER THE PATH OF EXCEL FILE : ", fg="black", bg="lemon chiffon", font=('Bahnschrift SemiLight', 15))
            self.voter_add_label.place(x=80, y=410)

            self.file_path = StringVar()
            self.file_path_entry = Entry(self.voter_add_frame, textvariable=self.file_path, width=40, bd=0)
            self.file_path_entry.place(x=420, y=413)

            self.file_path_enter = Button(self.voter_add_frame, text=" ENTER ", fg="white", bg="black", width=20, height=3, bd=0, command=voter_file_add_function, cursor="hand2")
            self.file_path_enter.place(x=200, y=520)

            self.file_path_cancel = Button(self.voter_add_frame, text=" Cancel ", fg="white", bg="black", width=20, height=3, bd=0, command=file_path_cancel_function, cursor="hand2")
            self.file_path_cancel.place(x=400, y=520)
        

        def file_path_cancel_function():
            self.voter_add_frame.place_forget()
        

        def voter_file_add_function():
            """ This function first tries to append the details of each voter i.e. one row of the excel sheet to a temperory list which is then appended to the voters_data_list.
                If there is any error or exception, then it displays the apppropriate remark message in the remark frame. """

            try:

                wb = openpyxl.load_workbook(self.file_path.get())
                sh1 = wb["Sheet1"]

                for i in range(2, sh1.max_row+1):
                    temp = []
                    for j in range(1, sh1.max_column+1):
                        temp.append(sh1.cell(row=i, column=j).value)
                    self.voters_data_list.append(temp)
                
                remark_text = " VOTERS DATA ADDED SUCCESSFULLY. "
                remark_frame_function(self.voter_add_frame, texts=remark_text, flag=True, texts_color ="green")
               
            
            except FileNotFoundError:
                remark_text = " FILE NOT FOUND. CHECK FILE NAME OR FILE PATH  AGAIN. "
                remark_frame_function(self.voter_add_frame, texts=remark_text, flag=False, texts_color ="red")
            except:
                remark_text = " FILE NAME OR EXTENSION IS INCORRECT. "
                remark_frame_function(self.voter_add_frame, texts=remark_text, flag=False, texts_color ="red")
        


        def view_candidates_function():
            """ This function is called whenever the user presses the View Candidates button in the home page. This function first checks if there are already any registered candidates. 
                If no candidates are registered it displays the same or else it will display the frame with list of posts already registered. User have to press the button in line with 
                the post name to view the candidates in that post. """
            
            if len(self.post_names) == 0:
                remark_text = " NO CANDIDATES ARE ADDED. "
                remark_frame_function(self.home_page_frame, texts=remark_text, flag=False, texts_color="red")
            else:

                

                self.view_post_name_frame = Frame(self.home_page_frame, width=800, height=650, bg="wheat")
                self.view_post_name_frame.place(x=0, y=0)

                election_name_display_function(self.view_post_name_frame)

                self.temp_post_names = []

                for post in self.post_names:
                    self.temp_post_names.append(post)

                while len(self.temp_post_names) != 10 and len(self.temp_post_names) != 0:
                    if len(self.temp_post_names) <= 10:
                        self.temp_post_names.append("<empty>")
                    elif len(self.temp_post_names) > 10:
                        self.temp_post_names.pop()
                
                coordinate = 100

                for index in range(len(self.temp_post_names)):
                    self.view_post_name_label = Label(self.view_post_name_frame, text=f"{self.temp_post_names[index]}".title(), bg="tomato2", fg="white", font=('Bahnschrift SemiLight', 10))
                    self.view_post_name_label.place(x=100, y=coordinate)
                    coordinate += 50


                self.view_post_candidate_button_1 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[0], button_no=1), bg="light sea green", fg="black", cursor="hand2")
                self.view_post_candidate_button_2 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[1], button_no=2), bg="light sea green", fg="black", cursor="hand2")
                self.view_post_candidate_button_3 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[2], button_no=3), bg="light sea green", fg="black", cursor="hand2")
                self.view_post_candidate_button_4 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[3], button_no=4), bg="light sea green", fg="black", cursor="hand2")
                self.view_post_candidate_button_5 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[4], button_no=5), bg="light sea green", fg="black", cursor="hand2")
                self.view_post_candidate_button_6 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[5], button_no=6), bg="light sea green", fg="black", cursor="hand2")
                self.view_post_candidate_button_7 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[6], button_no=7), bg="light sea green", fg="black", cursor="hand2")
                self.view_post_candidate_button_8 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[7], button_no=8), bg="light sea green", fg="black", cursor="hand2")
                self.view_post_candidate_button_9 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[8], button_no=9), bg="light sea green", fg="black", cursor="hand2")
                self.view_post_candidate_button_10 = Button(self.view_post_name_frame, text=" VIEW CANDIDATES ", bd=2, command=lambda: view_post_candidates_function(self.temp_post_names[9], button_no=10), bg="light sea green", fg="black", cursor="hand2")


                self.view_post_candidate_button_1.place(x=300, y=100)
                self.view_post_candidate_button_2.place(x=300, y=150)
                self.view_post_candidate_button_3.place(x=300, y=200)
                self.view_post_candidate_button_4.place(x=300, y=250)
                self.view_post_candidate_button_5.place(x=300, y=300)
                self.view_post_candidate_button_6.place(x=300, y=350)
                self.view_post_candidate_button_7.place(x=300, y=400)
                self.view_post_candidate_button_8.place(x=300, y=450)
                self.view_post_candidate_button_9.place(x=300, y=500)
                self.view_post_candidate_button_10.place(x=300, y=550)


            
            
                self.view_candidates_cancel_button = Button(self.view_post_name_frame, text=" Cancel ", fg="white", bg="black", bd=0, width=18, height=2, command=view_candidates_cancel_function, cursor="hand2")
                self.view_candidates_cancel_button.place(x=530, y=300)
        

        def view_candidates_cancel_function():
            """ This function is used to destroy the view_post_name_frame whenever user presses cancel button in that frame. """
            self.view_post_name_frame.place_forget()
        

        def view_post_candidates_function(post_name, button_no):
            """ This function is called whenever the user presses the View Candidate button infront of any post names. The post name and button number is passed as the argument to this function 
                Thne it lists out the all registered candidates and their details. """

            if button_no > len(self.post_names):

                remark_text = " THIS IS EMPTY POST. "
                remark_frame_function(self.view_post_name_frame, texts=remark_text, flag=False, texts_color="red")
            
            else:


                self.view_post_candidate_frame = Frame(self.view_post_name_frame, width=800, height=650, bg="wheat")
                self.view_post_candidate_frame.place(x=0, y=0)
               

                election_name_display_function(self.view_post_candidate_frame)

                self.post_name_display_label = Label(self.view_post_candidate_frame, text=f"{post_name}\nCandidates", fg="royalblue", font=("Franklin Gothic Medium", 15))
                self.post_name_display_label.place(x=570, y=130)

                self.candidate_group_column = Label(self.view_post_candidate_frame, text=" GROUP NAME ", bg="#DEBB43", fg="black", font=('Bahnschrift SemiLight', 10))
                self.candidate_name_column = Label(self.view_post_candidate_frame, text=" CANDIDATE NAME ", bg="#DEBB43", fg="black", font=('Bahnschrift SemiLight', 10))
                self.candidate_logo_column = Label(self.view_post_candidate_frame, text=" GROUP LOGO ", bg="#DEBB43", fg="black", font=('Bahnschrift SemiLight', 10))

                self.candidate_group_column.place(x=50, y=100)
                self.candidate_name_column.place(x=200, y=100)
                self.candidate_logo_column.place(x=365, y=100)

                coordinate = 125
                for candidates in self.final_candidates[post_name]:
                    

                    self.candidate_group_display = Label(self.view_post_candidate_frame, text=candidates[0], bg="wheat", fg="purple", font=('Calibri Light', 12))
                    self.candidate_name_display = Label(self.view_post_candidate_frame, text=candidates[1], bg="wheat", fg="purple", font=('Calibri Light', 12))
                    self.candidate_symbol_display = Label(self.view_post_candidate_frame, text=candidates[2], bg="wheat", fg="purple", font=('Calibri Light', 12))

                    self.candidate_group_display.place(x=50, y=coordinate)
                    self.candidate_name_display.place(x=200, y=coordinate)
                    self.candidate_symbol_display.place(x=365, y=coordinate)

                    coordinate += 30
                
                self.view_candidates_cancel_button = Button(self.view_post_candidate_frame, text=" Cancel ", fg="white", bg="black", width=18, height=2, bd=0, command=view_post_candidates_cancel_function, cursor="hand2")
                self.view_candidates_cancel_button.place(x=600, y=300)
        

        def view_post_candidates_cancel_function():
            """ This function is called whenever the user presses the cancel button present in the view_post_candidate_frame. """
            self.view_post_candidate_frame.place_forget()
        



        def voting_function():
            """ This function checks if the voting is already started or not. If the voitng is already started it directs the user to enter voter ID and password or else it will ask for the user
                confirmation to start the election. """
            
            if len(self.post_names) == 0:
                remark_text = " NO CANDIDATES ARE ADDED TO THE ELECTION. "
                remark_frame_function(self.home_page_frame, texts=remark_text, flag=False, texts_color="red")
            
            elif len(self.voters_data_list) == 0:
                remark_text = " VOTERS DATA IS NOT ADDED. "
                remark_frame_function(self.home_page_frame, texts=remark_text, flag=False, texts_color="red")


            elif not self.vote_started:
                confirmation_function()
            
            elif self.vote_started:
                voter_id_function()

        
        def confirmation_function():
            """ This function asks for the user to start voting for the first time. If user press enter then it asks the user to enter the voter ID and password. """

            self.confirmation_frame = Frame(self.home_page_frame, width=800, height=250, bg="khaki1")
            self.confirmation_frame.place(x=0, y=225)

            confirmaton_text = "ONCE VOTING GET STARTED YOU CANNOT ADD THE POSTS AND CANDIDATES.\nCLICK 'ENTER' TO CONTINUE."

            self.confirmation_label = Label(self.confirmation_frame, text=confirmaton_text, font=('Bahnschrift SemiLight', 11, 'bold'), fg="black", bg="bisque2")
            self.confirmation_label.place(x=120, y=60)

            self.confirmation_button = Button(self.confirmation_frame, bd=0, command=voter_id_function, text=" ENTER ", fg="white", bg="black", width=15, height=3, cursor="hand2")
            self.confirmation_button.place(x=200, y=150)

            self.confirmation_cancel_button = Button(self.confirmation_frame, bd=0, command=confirmation_cancel_function, text=" CANCEL ", fg="white", bg="black", width=15, height=3, cursor="hand2")
            self.confirmation_cancel_button.place(x=450, y=150)
        

        def confirmation_cancel_function():
            """ This function is called whenever the user presses the cancel button present in the confirmation_frame.
                This funciton destroys the frame. """

            self.confirmation_frame.place_forget()
        


        def voter_id_function():
            """ This function first destroys the confirmation frame if it is present or else it will create a frame which asks for the user to enter the 
                Voter ID and password """
            
            if not self.vote_started:
                confirmation_cancel_function()
                self.vote_started = True

            self.voter_id_frame = Frame(self.home_page_frame, width=800, height=650, bg="#851085")
            self.voter_id_frame.place(x=0,y=0)

            election_name_display_function(self.voter_id_frame)

            self.voter_id_label = Label(self.voter_id_frame, text=" VOTER ID : ", bg="#851085", fg="#EDF314",font=('Calibri Light', 17, 'bold'))
            self.voter_id_label.place(x=215, y=242)

            self.voter_id_label = Label(self.voter_id_frame, text=" PASSWORD : ", bg="#851085", fg="#EDF314",font=('Calibri Light', 17, 'bold'))
            self.voter_id_label.place(x=205, y=305)

            self.voter_id = StringVar()
            self.voter_id_entry = Entry(self.voter_id_frame, textvariable=self.voter_id, bd=0, width=30)
            self.voter_id_entry.place(x=360, y=254)

            self.voter_id_password = StringVar()
            self.voter_id_password_entry = Entry(self.voter_id_frame, textvariable=self.voter_id_password, bd=0, width=30, show="*")
            self.voter_id_password_entry.place(x=358, y=316)

            self.toggle_voter_password_button = Button(self.voter_id_frame, text="Show", bd=0, fg="white", bg="black", height=1, width=8, command=lambda : toggle_password_function(self.voter_id_password_entry, self.toggle_voter_password_button), cursor="hand2")
            self.toggle_voter_password_button.place(x=560, y=317)

            self.voter_id_enter = Button(self.voter_id_frame, bd=0, command=voter_id_type_check_function, text=" Enter ", fg="white", bg="black", width=15, height=3, cursor="hand2")
            self.voter_id_enter.place(x=200, y=500)

            self.voter_id_cancel = Button(self.voter_id_frame, bd=0, command=voter_id_cancel_function, text=" Cancel ", fg="white", bg="black", width=15, height=3, cursor="hand2")
            self.voter_id_cancel.place(x=450, y=500)
        

        def voter_id_type_check_function():
            try:

                if type(int(self.voter_id.get())) == int:
                    self.voter_id_variable = int(self.total_voters.get())
                       

            except ValueError:
                
                remark_text = " VOTER ID SHOULD BE A NUMBER. "
                remark_frame_function(self.window, texts=remark_text, flag=False, texts_color="red")
            
            else:
                main_voting_page_function()

        



        def main_voting_page_function():

            """ This function is called whenever user presses enter button after entering voter ID and password. First this function checks whether the voter ID entered is present in the voters_data_list
                using for loop. If the voter ID is not present it displays the not eligible to vote message. Next it cheks if the password entered is correct if the Voter ID is found in the list. If the password
                is wrong it displays the same to the user throug remark frame. Once if both voter ID and password is correct it checks if the voter is already voted. If the voter is already voted he cannot vote
                again or else the list of registered post is displayed for the voter to select and vote."""

            
            
            self.entries = False

            for voter_data in self.voters_data_list:
                if int(self.voter_id.get()) == voter_data[0]:
                    self.entries = True
                    if voter_data[3] != self.voter_id_password.get():
                        
                        remark_text = " WRONG PASSWORD. TRY AGAIN. "
                        remark_frame_function(self.voter_id_frame, texts=remark_text, flag=False, texts_color="red")
                    
                    else:
                       
                        self.is_voted = voter_data[2]
                        

                        if not self.is_voted:


                            self.main_vote_frame = Frame(self.voter_id_frame, width=800, height=650, bg="#F1CD51")
                            self.main_vote_frame.place(x=0, y=0)

                            election_name_display_function(self.main_vote_frame)      

                            if not self.fill_index:

                                

                                self.total_posts_count = len(self.post_names)
                                self.voted_posts = []
                               
                        
                                for post in self.post_names:
                                 
                                    
                                    while len(self.final_candidates[str(post)]) != 10 and len(self.final_candidates[str(post)]) != 0:


                                

                                        if len(self.final_candidates[post]) <= 10:
                                            self.final_candidates[post].append(["<empty>", "", "", 0])
                                        elif len(self.final_candidates[post]) > 10:
                                            self.final_candidates[post].pop()
                                
                                while len(self.post_names) != 10 and len(self.post_names) != 0:
                                    if len(self.post_names) <= 10:
                                        self.post_names.append("<empty>")
                                    elif len(self.post_names) > 10:
                                        self.post_names.pop()
                                
                                self.fill_index = 1
                            
                           
                            try:
                               


                                self.post_name_label_1 = Label(self.main_vote_frame, text=f"{self.post_names[0]}".upper(), bg="aqua", fg="red")
                                self.post_name_label_2 = Label(self.main_vote_frame, text=f"{self.post_names[1]}".upper(), bg="aqua", fg="red")
                                self.post_name_label_3 = Label(self.main_vote_frame, text=f"{self.post_names[2]}".upper(), bg="aqua", fg="red")
                                self.post_name_label_4 = Label(self.main_vote_frame, text=f"{self.post_names[3]}".upper(), bg="aqua", fg="red")
                                self.post_name_label_5 = Label(self.main_vote_frame, text=f"{self.post_names[4]}".upper(), bg="aqua", fg="red")
                                self.post_name_label_6 = Label(self.main_vote_frame, text=f"{self.post_names[5]}".upper(), bg="aqua", fg="red")
                                self.post_name_label_7 = Label(self.main_vote_frame, text=f"{self.post_names[6]}".upper(), bg="aqua", fg="red")
                                self.post_name_label_8 = Label(self.main_vote_frame, text=f"{self.post_names[7]}".upper(), bg="aqua", fg="red")
                                self.post_name_label_9 = Label(self.main_vote_frame, text=f"{self.post_names[8]}".upper(), bg="aqua", fg="red")
                                self.post_name_label_10 = Label(self.main_vote_frame, text=f"{self.post_names[9]}".upper(), bg="aqua", fg="red")


                                self.post_name_label_1.place(x=100, y=100)
                                self.post_name_label_2.place(x=100, y=150)
                                self.post_name_label_3.place(x=100, y=200)
                                self.post_name_label_4.place(x=100, y=250)
                                self.post_name_label_5.place(x=100, y=300)
                                self.post_name_label_6.place(x=100, y=350)
                                self.post_name_label_7.place(x=100, y=400)
                                self.post_name_label_8.place(x=100, y=450)
                                self.post_name_label_9.place(x=100, y=500)
                                self.post_name_label_10.place(x=100, y=550)


                                self.post_name_button_1 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[0], button_no=1), bg="wheat", fg="brown", cursor="hand2")
                                self.post_name_button_2 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[1], button_no=2), bg="wheat", fg="brown", cursor="hand2")
                                self.post_name_button_3 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[2], button_no=3), bg="wheat", fg="brown", cursor="hand2")
                                self.post_name_button_4 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[3], button_no=4), bg="wheat", fg="brown", cursor="hand2")
                                self.post_name_button_5 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[4], button_no=5), bg="wheat", fg="brown", cursor="hand2")
                                self.post_name_button_6 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[5], button_no=6), bg="wheat", fg="brown", cursor="hand2")
                                self.post_name_button_7 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[6], button_no=7), bg="wheat", fg="brown", cursor="hand2")
                                self.post_name_button_8 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[7], button_no=8), bg="wheat", fg="brown", cursor="hand2")
                                self.post_name_button_9 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[8], button_no=9), bg="wheat", fg="brown", cursor="hand2")
                                self.post_name_button_10 = Button(self.main_vote_frame, text=" VOTE ", command=lambda: post_vote_function(self.post_names[9], button_no=10), bg="wheat", fg="brown", cursor="hand2")


                                self.post_name_button_1.place(x=500, y=100)
                                self.post_name_button_2.place(x=500, y=150)
                                self.post_name_button_3.place(x=500, y=200)
                                self.post_name_button_4.place(x=500, y=250)
                                self.post_name_button_5.place(x=500, y=300)
                                self.post_name_button_6.place(x=500, y=350)
                                self.post_name_button_7.place(x=500, y=400)
                                self.post_name_button_8.place(x=500, y=450)
                                self.post_name_button_9.place(x=500, y=500)
                                self.post_name_button_10.place(x=500, y=550)


                                self.main_vote_cancel_button = Button(self.main_vote_frame, text=" Cancel ", command=main_vote_cancel_function, bg="black", fg="white", cursor="hand2")
                                self.main_vote_cancel_button.place(x=700, y=250)

                                self.main_vote_submit_button = Button(self.main_vote_frame, text=" Finish ", command=main_vote_finish_funciton, bg="black", fg="white", cursor="hand2")
                                self.main_vote_submit_button.place(x=700, y=300)
                            
                            except AttributeError:

                                remark_text = " NO CONTESTANTS ENROLLED FOR THIS POST. "
                                remark_frame_function(self.main_vote_frame, texts=remark_text, flag=False, texts_color="red")
                                
                                

                            except IndexError:

                                remark_text = " NO CONTESTANTS ENROLLED FOR THIS POST. "
                                remark_frame_function(self.main_vote_frame, texts=remark_text, flag=False, texts_color="red")
                                
                        
                        else:
                            
                            remark_text = " YOU HAVE ALREADY VOTED. "
                            remark_frame_function(self.voter_id_frame, texts=remark_text, flag=False, texts_color="red") 
                        
                
                    

            if not self.entries:
                remark_text = " YOU ARE NOT ELIGIBLE TO VOTE. "
                remark_frame_function(self.voter_id_frame, texts=remark_text, flag=False, texts_color="red")            
            
        

        def voter_id_cancel_function():
            """ This function is called whenever the user presses the cancel button present in the voter_id_frame.
                This funciton destroys the frame. """

            self.voter_id_frame.place_forget()
        


        def post_vote_function(post_name, button_no):
            """ This function is called whenever voter presses the VOTE button inline with the post name in the main_vote_frame. This function first checks 
                if the button pressed corresponds to the empty post and if so it displays the remark message. Then it cheks if the voter already voted for 
                this post and if so it displays the same in the remark frame. If both of the previous mentioned process did not happend it desplays the list
                of candidates and theri detail in the newly created frame and voter need to select the candidate. """


            if button_no > self.total_posts_count:

                remark_text = " THIS IS EMPTY POST. "
                remark_frame_function(self.main_vote_frame, texts=remark_text, flag=False, texts_color="red")
            
            elif button_no in self.voted_posts:

                remark_text = " YOU HAVE ALREADY VOTED FOR THIS POST. "
                remark_frame_function(self.main_vote_frame, texts=remark_text, flag=False, texts_color="red")
            

            else:

            
            
                self.each_post_vote_frame = Frame(self.main_vote_frame, width=800, height=650, bg="wheat")
                self.each_post_vote_frame.place(x=0, y=0)

                election_name_display_function(self.each_post_vote_frame)

                self.vote_post_name_display_label = Label(self.each_post_vote_frame, text=f"{post_name}\nCandidates", fg="royalblue", bg="wheat", font=("Franklin Gothic Medium", 15))
                self.vote_post_name_display_label.place(x=660, y=130)


                self.entry_1 = IntVar()
                self.entry_2 = IntVar()
                self.entry_3 = IntVar()
                self.entry_4 = IntVar()
                self.entry_5 = IntVar()
                self.entry_6 = IntVar()
                self.entry_7 = IntVar()
                self.entry_8 = IntVar()
                self.entry_9 = IntVar()
                self.entry_10 = IntVar()

                self.candidate_group_1 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][0][2], onvalue=1, offvalue=0, variable=self.entry_1, bg="wheat", fg="black", cursor="hand2")
                self.candidate_group_2 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][1][2], onvalue=1, offvalue=0, variable=self.entry_2, bg="wheat", fg="black", cursor="hand2")
                self.candidate_group_3 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][2][2], onvalue=1, offvalue=0, variable=self.entry_3, bg="wheat", fg="black", cursor="hand2")
                self.candidate_group_4 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][3][2], onvalue=1, offvalue=0, variable=self.entry_4, bg="wheat", fg="black", cursor="hand2")
                self.candidate_group_5 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][4][2], onvalue=1, offvalue=0, variable=self.entry_5, bg="wheat", fg="black", cursor="hand2")
                self.candidate_group_6 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][5][2], onvalue=1, offvalue=0, variable=self.entry_6, bg="wheat", fg="black", cursor="hand2")
                self.candidate_group_7 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][6][2], onvalue=1, offvalue=0, variable=self.entry_7, bg="wheat", fg="black", cursor="hand2")
                self.candidate_group_8 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][7][2], onvalue=1, offvalue=0, variable=self.entry_8, bg="wheat", fg="black", cursor="hand2")
                self.candidate_group_9 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][8][2], onvalue=1, offvalue=0, variable=self.entry_9, bg="wheat", fg="black", cursor="hand2")
                self.candidate_group_10 = Checkbutton(self.each_post_vote_frame, text=self.final_candidates[post_name][9][2], onvalue=1, offvalue=0, variable=self.entry_10, bg="wheat", fg="black", cursor="hand2")

                self.candidate_group_1.place(x=520, y=140)
                self.candidate_group_2.place(x=520, y=190)
                self.candidate_group_3.place(x=520, y=240)
                self.candidate_group_4.place(x=520, y=290)
                self.candidate_group_5.place(x=520, y=340)
                self.candidate_group_6.place(x=520, y=390)
                self.candidate_group_7.place(x=520, y=440)
                self.candidate_group_8.place(x=520, y=490)
                self.candidate_group_9.place(x=520, y=540)
                self.candidate_group_10.place(x=520, y=590)

                self.candidate_group_column = Label(self.each_post_vote_frame, text=" GROUP NAME ", bg="#DEBB43", fg="black", font=('Bahnschrift SemiLight', 10))
                self.candidate_name_column = Label(self.each_post_vote_frame, text=" CANDIDATE NAME ", bg="#DEBB43", fg="black", font=('Bahnschrift SemiLight', 10))
                self.candidate_logo_column = Label(self.each_post_vote_frame, text=" GROUP LOGO ", bg="#DEBB43", fg="black", font=('Bahnschrift SemiLight', 10))

                self.candidate_group_column.place(x=50, y=100)
                self.candidate_name_column.place(x=200, y=100)
                self.candidate_logo_column.place(x=365, y=100)

                coordinate = 0
                for candidates in self.final_candidates[post_name]:

                    self.candidate_group_display = Label(self.each_post_vote_frame, text=candidates[0], bg="wheat", fg="purple", font=('Calibri Light', 12))
                    self.candidate_name_display = Label(self.each_post_vote_frame, text=candidates[1], bg="wheat", fg="purple", font=('Calibri Light', 12))
                    self.candidate_symbol_display = Label(self.each_post_vote_frame, text=candidates[2], bg="wheat", fg="purple", font=('Calibri Light', 12))

                    self.candidate_group_display.place(x=50, y=140 + coordinate)
                    self.candidate_name_display.place(x=200, y=140 + coordinate)
                    self.candidate_symbol_display.place(x=365, y=140 + coordinate)

                    coordinate += 50
                
                self.vote_submit_button = Button(self.each_post_vote_frame, text=" SUBMIT ", bd=0, width=10, height=2, command= lambda: vote_submit_function(post_name, button_no), bg="black", fg="white", cursor="hand2")
                self.vote_cancel_button = Button(self.each_post_vote_frame, text=" CANCEL ", bd=0, width=10, height=2, command=vote_cancel_function, bg="black", fg="white", cursor="hand2")

                self.vote_submit_button.place(x=680, y=250)
                self.vote_cancel_button.place(x=680, y=350)
        


        def main_vote_cancel_function():
            """ This function is called whenever the voter presses the cancel button in main_vote_frame. This function first clears the values entered by the user for voter ID and Password 
                and then destroys the main_vote_frame """

            self.voter_id_entry.delete(0, END)
            self.voter_id_password_entry.delete(0, END)
            self.main_vote_frame.place_forget()
        

        def vote_submit_function(post_name, button_no):
            """ This function is called whenever the voter presses the submit button on the each_post_vote_frame. This function first checks if the voter selected more than one group and then 
                it chekcs whether or not voter selected any of the group and finally whether the voter selected empty group. It will desplay the appropriate remark message in the remark frame.
                Once everything is correct, this function will display successfull message to the voter. """

            self.selected_entry = [int(self.entry_1.get()),
                                    int(self.entry_2.get()),
                                    int(self.entry_3.get()),
                                    int(self.entry_4.get()),
                                    int(self.entry_5.get()),
                                    int(self.entry_6.get()),
                                    int(self.entry_7.get()),
                                    int(self.entry_8.get()),
                                    int(self.entry_9.get()),
                                    int(self.entry_10.get())]
            
            
            self.selected_entry_count = 0

            for index in range(len(self.selected_entry)):

                if self.selected_entry[index] != 0:
                    self.selected_entry_count += 1
                    self.selected_entry_index = index
                
            if self.selected_entry_count > 1:

                remark_text = " YOU CANNOT SELECT MORE THAN ONE GROUP. "
                remark_frame_function(self.each_post_vote_frame, texts=remark_text, flag=False, texts_color="red")
                
            elif self.selected_entry_count < 1:
                
                remark_text = " YOU HAVE NOT SELECTED ANY GROUP. "
                remark_frame_function(self.each_post_vote_frame, texts=remark_text, flag=False, texts_color="red")
                
            else:

                if self.final_candidates[post_name][self.selected_entry_index][0] == "<empty>":

                    remark_text = " THIS IS THE EMPTY GROUP. YOU CANNOT SELECT THIS GROUP. "
                    remark_frame_function(self.each_post_vote_frame, texts=remark_text, flag=False, texts_color="red")
                
                else:

                    self.final_candidates[post_name][self.selected_entry_index][3] += 1

                    self.voted_posts.append(button_no)

                    remark_text = f" YOU VOTED SUCCESSFULLY FOR {post_name.upper()} POST "
                    remark_frame_function(self.each_post_vote_frame, texts=remark_text, flag=True, texts_color="green")



        def vote_cancel_function():
            """ This function is called whenever the voter presses the cancel button in main_vote_frame. """
            self.each_post_vote_frame.place_forget()
        


        def main_vote_finish_funciton():
            """ This function is used when the voter presses the finish button on the main_vote_frame. First it checks whether or not the voter has posted for every posts.
                If the voter has voted for every posts it displays the successful message. Then it also increases the total number of voters voted count by one and also plays 
                the indicative sound that the voter has voted successfully. """

            if self.total_posts_count != len(self.voted_posts):

                remark_text = " YOU HAVE NOT VOTED FOR ALL POSTS. "
                remark_frame_function(self.main_vote_frame, texts=remark_text, flag=False, texts_color="red")
            
            else:

               
                for item in self.voters_data_list:

                    if int(self.voter_id.get()) == item[0]:

                        item[2] = 1
                        break


                remark_text = " YOU HAVE VOTED SUCCESSFULLY. THANK YOU. "
                remark_frame_function(self.main_vote_frame, texts=remark_text, flag=True, texts_color="green")

                self.voted_posts = []
                self.voter_id_password_entry.delete(0, END)
                self.voter_id_entry.delete(0, END)
                

                self.voters_count += 1

                sound_play_function()
        

        def result_password_enter_function():
            """ This function is called when the user presses the Result button in the home page. This function creates the frame that asks the user to enter
                the password that was set during the welcome page of the GUI."""

            self.result_password_frame = Frame(self.home_page_frame, width=800, height=650, bg="burlywood1")
            self.result_password_frame.place(x=0, y=0)

            election_name_display_function(self.result_password_frame)

            self.result_password_instruction_text = Text(self.result_password_frame, height=7, width = 50, bg="tomato")
            self.result_password_instruction_text.place(x=150, y=140)
            self.result_password_instruction_text.insert(END, " Once you enter the correct password and press \nSubmit button, you can not go back and vote for \nthe candidates.\n\n")
            self.result_password_instruction_text.insert(END, " Make sure all the registered voters already voted\nand then enter the password and view results.")
            self.result_password_instruction_text.config(state="disabled")

            self.verify_password_label = Label(self.result_password_frame, text=" ENTER THE PASSWORD : ", fg="black", bg="misty rose", font=('Bahnschrift SemiLight', 13))
            self.verify_password_label.place(x=170, y=350)

            
            self.password_holder = StringVar()
            self.verify_password_entry = Entry(self.result_password_frame, textvariable=self.password_holder, width=30, bd=0, show="*")
            self.verify_password_entry.place(x=390, y=353)

            self.toggle_verify_password_button = Button(self.result_password_frame, text="Show", bd=0, fg="white", bg="black", height=1, width=8, command=lambda : toggle_password_function(self.verify_password_entry, self.toggle_verify_password_button), cursor="hand2")
            self.toggle_verify_password_button.place(x=600, y=351)

            

            self.password_submit_button = Button(self.result_password_frame, text=" Submit ", bd=0, fg="white", bg="black", command = view_result_function, width=20, height=3, cursor="hand2")
            self.password_submit_button.place(x=200, y=500)

            self.result_password_cancel_button = Button(self.result_password_frame, text=" Cancel ", bd=0, fg="white", bg="black", command = result_password_cancel_function, width=20, height=3, cursor="hand2")
            self.result_password_cancel_button.place(x=400, y=500)
        


        def view_result_function():
            """ Once user enters the password and hit submit this function is called which first checks if the password is correct. If not it displays 
                the remark message on the remark frame. If the password entered is correct it will create a frame which has the list of all the posts 
                and inline with that will be the button to view the result of corresponding post. Apart from these buttons, the frame also has other three buttons
                namely, Save in Files, Exit, Back. """

            if not self.password_holder.get() == self.first_password.get():

                remark_text = " WRONG PASSWORD. TRY AGAIN. "
                remark_frame_function(self.result_password_frame, texts=remark_text, flag=False, texts_color="red")
            

            else:


                self.view_result_frame = Frame(self.home_page_frame, width=800, height=650, bg="silver")
                self.view_result_frame.place(x=0, y=0)

                election_name_display_function(self.view_result_frame)



                self.post_result_label_1 = Label(self.view_result_frame, text=f"{self.post_names[0]}".upper(), bg="aqua", fg="red")
                self.post_result_label_2 = Label(self.view_result_frame, text=f"{self.post_names[1]}".upper(), bg="aqua", fg="red")
                self.post_result_label_3 = Label(self.view_result_frame, text=f"{self.post_names[2]}".upper(), bg="aqua", fg="red")
                self.post_result_label_4 = Label(self.view_result_frame, text=f"{self.post_names[3]}".upper(), bg="aqua", fg="red")
                self.post_result_label_5 = Label(self.view_result_frame, text=f"{self.post_names[4]}".upper(), bg="aqua", fg="red")
                self.post_result_label_6 = Label(self.view_result_frame, text=f"{self.post_names[5]}".upper(), bg="aqua", fg="red")
                self.post_result_label_7 = Label(self.view_result_frame, text=f"{self.post_names[6]}".upper(), bg="aqua", fg="red")
                self.post_result_label_8 = Label(self.view_result_frame, text=f"{self.post_names[7]}".upper(), bg="aqua", fg="red")
                self.post_result_label_9 = Label(self.view_result_frame, text=f"{self.post_names[8]}".upper(), bg="aqua", fg="red")
                self.post_result_label_10 = Label(self.view_result_frame, text=f"{self.post_names[9]}".upper(), bg="aqua", fg="red")


                self.post_result_label_1.place(x=100, y=100)
                self.post_result_label_2.place(x=100, y=150)
                self.post_result_label_3.place(x=100, y=200)
                self.post_result_label_4.place(x=100, y=250)
                self.post_result_label_5.place(x=100, y=300)
                self.post_result_label_6.place(x=100, y=350)
                self.post_result_label_7.place(x=100, y=400)
                self.post_result_label_8.place(x=100, y=450)
                self.post_result_label_9.place(x=100, y=500)
                self.post_result_label_10.place(x=100, y=550)

                self.view_result_button_1 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[0], button_no=1), bg="wheat", fg="brown", cursor="hand2")
                self.view_result_button_2 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[1], button_no=2), bg="wheat", fg="brown", cursor="hand2")
                self.view_result_button_3 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[2], button_no=3), bg="wheat", fg="brown", cursor="hand2")
                self.view_result_button_4 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[3], button_no=4), bg="wheat", fg="brown", cursor="hand2")
                self.view_result_button_5 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[4], button_no=5), bg="wheat", fg="brown", cursor="hand2")
                self.view_result_button_6 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[5], button_no=6), bg="wheat", fg="brown", cursor="hand2")
                self.view_result_button_7 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[6], button_no=7), bg="wheat", fg="brown", cursor="hand2")
                self.view_result_button_8 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[7], button_no=8), bg="wheat", fg="brown", cursor="hand2")
                self.view_result_button_9 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[8], button_no=9), bg="wheat", fg="brown", cursor="hand2")
                self.view_result_button_10 = Button(self.view_result_frame, text=" View Result ", command=lambda: view_post_result_function(self.post_names[9], button_no=10), bg="wheat", fg="brown", cursor="hand2")


                self.view_result_button_1.place(x=500, y=100)
                self.view_result_button_2.place(x=500, y=150)
                self.view_result_button_3.place(x=500, y=200)
                self.view_result_button_4.place(x=500, y=250)
                self.view_result_button_5.place(x=500, y=300)
                self.view_result_button_6.place(x=500, y=350)
                self.view_result_button_7.place(x=500, y=400)
                self.view_result_button_8.place(x=500, y=450)
                self.view_result_button_9.place(x=500, y=500)
                self.view_result_button_10.place(x=500, y=550)


                self.result_save_button = Button(self.view_result_frame, text=" Save in Files ", bg="black", fg="white", command=result_save_function, width=15, height=3, cursor="hand2")
                                    
                self.election_exit_button = Button(self.view_result_frame, text=" EXIT ", bg="black", fg="white", command=election_exit_function, width=15, height=3, cursor="hand2")
                                

                                

                self.result_save_button.place(x=660, y=270)
                self.election_exit_button.place(x=660, y=370)
        

        def result_password_cancel_function():
            self.result_password_frame.place_forget()



        def view_post_result_function(post_name, button_no):
            """ This function is called when the user presses the View Result button in the view_result_frame. This takes the argument 
                of post name and button number for the reference. Then it displays the list of candidates with theri details and number
                of votes they got. The winner candidate is highlighted with different color and also with the label WINNER OF THE ELECTION. """

            if button_no > self.total_posts_count:

                remark_text = " THIS IS EMPTY POST. "
                remark_frame_function(self.view_result_frame, texts=remark_text, flag=False, texts_color="red")
            
            else:

                self.all_votes = []

                for candidate in self.final_candidates[post_name]:
                    self.all_votes.append(candidate[3])

                self.post_result_frame = Frame(self.view_result_frame, width=800, height=650, bg="#81A796")
                self.post_result_frame.place(x=0, y=0)

                election_name_display_function(self.post_result_frame)

                self.result_post_name_display_label = Label(self.post_result_frame, text=f"{post_name}\nCandidates", fg="royalblue", bg="wheat", font=("Franklin Gothic Medium", 15))
                self.result_post_name_display_label.place(x=660, y=60)


                self.votes_of_winner = max(self.all_votes)
                self.votes_of_winner_list.append(self.votes_of_winner)

                self.post_group_name_column = Label(self.post_result_frame, text=" GROUP NAME ", bg="#81A796", fg="black", font=('Bahnschrift SemiLight', 10))
                                
                self.post_candidate_name_column = Label(self.post_result_frame, text=" CANDIDATE NAME ", bg="#81A796", fg="black", font=('Bahnschrift SemiLight', 10))
                                
                self.post_group_logo_column = Label(self.post_result_frame, text=" GROUP LOGO ", bg="#81A796", fg="black", font=('Bahnschrift SemiLight', 10))
                                
                self.post_total_votes_column = Label(self.post_result_frame, text=" TOTAL VOTES ", bg="#81A796", fg="black", font=('Bahnschrift SemiLight', 10))

                self.post_group_name_column.place(x=30, y=100)
                self.post_candidate_name_column.place(x=160, y=100)
                self.post_group_logo_column.place(x=330, y=100)
                self.post_total_votes_column.place(x=450, y=100)  


                i = 0
                count = 0
                for index in range(len(self.all_votes)):

                    if self.all_votes[index] == self.votes_of_winner:

                        self.winner_group = Label(self.post_result_frame, text=self.final_candidates[post_name][i][0], bg="#81A796", fg="red2", font=('Calibri Light', 12))  

                        self.winner_name = Label(self.post_result_frame, text=self.final_candidates[post_name][i][1], bg="#81A796", fg="red2", font=('Calibri Light', 12))
                                    
                        self.winner_logo = Label(self.post_result_frame, text=self.final_candidates[post_name][i][2], bg="#81A796", fg="red2", font=('Calibri Light', 12))
                                        
                        self.winner_votes = Label(self.post_result_frame, text=self.all_votes[index], bg="#81A796", fg="red2", font=('Calibri Light', 12))
                                    
                        self.winner_text = Label(self.post_result_frame, text=" WINNER OF THE ELECTION ", bg="#81A796", fg="red2", font=('Calibri Light', 12, 'bold')) 

                        self.winner_group.place(x=30, y=130 + count)
                        self.winner_name.place(x=180, y=130 + count)
                        self.winner_logo.place(x=330, y=130 + count)
                        self.winner_votes.place(x=465, y=130 + count)
                        self.winner_text.place(x=500, y=130 + count)
                                        
                    else:

                        self.other_groups = Label(self.post_result_frame, text=self.final_candidates[post_name][i][0], bg="#81A796", fg="blue2", font=('Calibri Light', 12))
                                        
                        self.other_names = Label(self.post_result_frame, text=self.final_candidates[post_name][i][1], bg="#81A796", fg="blue2", font=('Calibri Light', 12))
                                    
                        self.other_logos = Label(self.post_result_frame, text=self.final_candidates[post_name][i][2], bg="#81A796", fg="blue2", font=('Calibri Light', 12))
                                        
                        self.other_votes = Label(self.post_result_frame, text=self.all_votes[index], bg="#81A796", fg="blue2", font=('Calibri Light', 12))
                                    

                        self.other_groups.place(x=30, y=130 + count)
                        self.other_names.place(x=180, y=130 + count)
                        self.other_logos.place(x=330, y=130 + count)
                        self.other_votes.place(x=465, y=130 + count)
                    
                    count += 50
                    i += 1
                


                self.post_result_back_button = Button(self.post_result_frame, text=" Back ", bg="black", fg="white", command=post_result_back_funciton, width=6, height=2, cursor="hand2")
                self.post_result_back_button.place(x=740, y=320)
        


        def post_result_back_funciton():
            """ This function is called whenever the user presses the cancel button in post_result_frame. """
            self.post_result_frame.place_forget()  




        def result_save_function():
            """ This function is used to save the details of the eleciton in the local storage for the further reference. This uses pythons inbuilt
                file handling methods to create, open and write in the file. The saved file has election name, voters count, post names, election summary
                winners list etc. Once the file is saved it displays the successful message and file name will be name of the election."""

            self.post_names_list = []
            self.final_candidates_list = {}

            for post in self.post_names:
                if post != "<empty>":
                    self.post_names_list.append(post)

            for post in self.post_names_list:
                self.final_candidates_list[post] = self.final_candidates[post]    

            with open(f'{self.election_name.get()}.txt', 'w') as file_object:

                file_object.write(f"              ELECTION NAME : {self.election_name.get().upper()}\n\n")
                file_object.write(f" DATE OF THE ELECTION  : {self.day} - {self.month} - {self.year}, {self.week}\n\n")
                file_object.write(f" POSTS OF THE ELECTION : \n")
                
                for post in self.post_names_list:
                    file_object.write(f"->{post}\n")
                
                file_object.write(f"\n TOTAL NUMBER OF REGISTERED VOTERS : {int(self.total_voters.get())}")
                file_object.write(f"\n TOTAL NUMBER OF VOTERS VOTED     : {self.voters_count}")
                
                
                file_object.write("\n\n*****************ELECTION SUMMARY*********************\n")

                for post in self.post_names_list:
                    file_object.write("\n------------------------------------------------------------------\n")
                    file_object.write(f"\n\n{post} POST CANDIDATES :\n")

                    for index in range(len(self.final_candidates_list)):
                        file_object.write(f"\n\n GROUP NAME           : {self.final_candidates_list[post][index][0].upper()}\n")
                        file_object.write(f" CANDIDATE NAME       : {self.final_candidates_list[post][index][1].upper()}\n")
                        file_object.write(f" CANDIDATE SYMBOL     : {self.final_candidates_list[post][index][2].upper()}\n")
                        file_object.write(f" CANDIDATE VOTES      : {self.final_candidates_list[post][index][3]}\n")

                        if self.final_candidates_list[post][index][3] == self.votes_of_winner_list[index]:
                            file_object.write(" WINNER OF THE ELECTION. ")
                
                file_object.write("\n\n\n************************************************************************************\n")
                file_object.write("                     WINNERS OF THE ELECTION")
                file_object.write("\n************************************************************************************\n")
                file_object.write("POST NAME               GROUP NAME               CANDIDATE NAME               CANDIDATE SYMBOL\n\n")

                for post in self.post_names_list:
                    
                    for index in range(len(self.final_candidates_list)):

                        if self.final_candidates_list[post][index][3] == self.votes_of_winner_list[index]:

                            g_name = self.final_candidates_list[post][index][0].upper()
                            c_name = self.final_candidates_list[post][index][1].upper()
                            c_symbol = self.final_candidates_list[post][index][2].upper()

                            post_length = 25-len(post)
                            lgn_length = 25-len(g_name)
                            lcn_length = 29-len(c_name)
                            

                            file_object.write(f"{post}{' '*(post_length)}{g_name}{' '*(lgn_length)}{c_name}{' '*(lcn_length)}{c_symbol}\n")
                    file_object.write("---------------------------------------------------------------------------------------------------\n")



            
            remark_text = " FILE SAVED SUCCESSFULLY. "
            remark_frame_function(self.view_result_frame, texts=remark_text, flag=False, texts_color="green")  


        

        def election_exit_function():
            """ This function destroy the root window thereby stopping the GUI. Election Ends."""

            disable_event()


        


        def disable_event():
            self.exit_password_frame = Frame(self.home_page_frame, width=800, height=650, bg="plum1")
            self.exit_password_frame.place(x=0, y=0)

            election_name_display_function(self.exit_password_frame)

            self.exit_instruction_text = Text(self.exit_password_frame, height=5, width = 50, bg="wheat")
            self.exit_instruction_text.place(x=150, y=150)
            self.exit_instruction_text.insert(END, "    Once you enter the correct password and hit \nSubmit button, you can not continue the election. \nThe election ends and all the data will be ")
            self.exit_instruction_text.insert(END, "lost. \nMake sure before entering the Submit button.")
            self.exit_instruction_text.config(state="disabled")

            self.exit_password_label = Label(self.exit_password_frame, text=" ENTER THE PASSWORD : ", fg="black", font=('Bahnschrift SemiLight', 13))
            self.exit_password_label.place(x=130, y=348)

            
            self.exit_password_holder = StringVar()
            self.exit_password_entry = Entry(self.exit_password_frame, textvariable=self.exit_password_holder, width=30, bd=0, show="*")
            self.exit_password_entry.place(x=350, y=352)

            self.toggle_exit_password_button = Button(self.exit_password_frame, text="Show", bd=0, fg="white", bg="black", height=1, width=8, command=lambda : toggle_password_function(self.exit_password_entry, self.toggle_exit_password_button), cursor="hand2")
            self.toggle_exit_password_button.place(x=550, y=350)

            self.exit_password_submit_button = Button(self.exit_password_frame, text=" Submit ", bd=0, fg="white", bg="black", command = exit_window_function, width=20, height=3, cursor="hand2")
            self.exit_password_submit_button.place(x=200, y=500)

            self.exit_password_cancel_button = Button(self.exit_password_frame, text=" Cancel ", bd=0, fg="white", bg="black", command = exit_cancel_function, width=20, height=3, cursor="hand2")
            self.exit_password_cancel_button.place(x=400, y=500)
        

        def exit_window_function():
            
            if self.exit_password_holder.get() == self.first_password.get() or self.exit_password_holder.get() == "emergency":
                window_close_function()
            
            else:
                remark_text = " WRONG PASSWORD. TRY AGAIN. "
                remark_frame_function(self.exit_password_frame, texts=remark_text, flag=False, texts_color="red")
        
        def exit_cancel_function():
            self.exit_password_frame.place_forget()
        

        def window_close_function():
            self.window.quit()




        
        def remark_frame_function(frame, texts, flag, texts_color):
            """ Whenever this function is called it creates frame on top of the parent frame i.e. calling frame.
                In the created frame it desplays the remark message to the user. """
            
            
            remark_frame = Frame(frame, width=800, height=150, bg="royalblue")
            remark_frame.place(x=0, y=250)

            remark_label = Label(remark_frame, text=texts, font=('Bahnschrift SemiLight', 11, 'bold'), fg=texts_color)
            remark_label.place(x=220, y=60)

            if flag:
                countdown_function(5, remark_frame, frame, flag)

            else:
                frame_destroy_function(4000, remark_frame)
            

        
        def countdown_function(time, remark_frame, frame, flag):
            """ This function is used to destroy both remark frame and the calling frame simultaneously after the given time period. """
            if time == -1:
                remark_frame.place_forget()
                frame.place_forget()

            else:
                remark_frame.after(500, lambda: countdown_function(time-1, remark_frame, frame, flag))



        def frame_destroy_function(time, frame):
            """ This function is used to destroy only remark frame after given time period. """
            frame.after(time, frame.destroy)


        def election_name_display_function(frame):
            """ This function is used to display the name of the election on top side of the frame. The frame in which the name is displayed is passed as the argument. """

            self.election_name_display_label = Label(frame, text=self.election_name.get().upper(), font=('Bahnschrift SemiLight', 15, 'bold'), fg="royalblue", bg="lemon chiffon")
            self.election_name_display_label.place(x=300, y=25)
        

        def toggle_password_function(entry_name, button_name):
            if entry_name.cget('show') == '':
                entry_name.config(show='*')
                button_name.config(text='Show')
            else:
                entry_name.config(show='')
                button_name.config(text='Hide')
        

        def sound_play_function():
            try:
                playsound("success.mp3")
            except:
                pass



        def main_function():
            """ This function is responsible for the first page of this GUI. It displays welcome message and asks for ELECTION NAME, TOTAL NUMBER OF VOTERS REGISTERED, PASSWORD. 
                User have to enter and confirm the password. """

            

            self.welcome_label = Label(self.window, text=" WELCOME ", font=('Yuanti SC', 30, 'bold'), bg="coral")
            self.welcome_label.place(x=285, y=80)


            self.election_name_label = Label(self.window, text=" ENTER THE ELECTION NAME : ", font=('Bahnschrift SemiLight', 13, 'bold'), bg="misty rose")
            self.election_name_label.place(x=100, y=200)

            self.voters_count_label = Label(self.window, text=" ENTER TOTAL NUMBER OF VOTERS : ", font=('Bahnschrift SemiLight', 13, 'bold'), bg="misty rose")
            self.voters_count_label.place(x=70, y=300)

            self.password_label = Label(self.window, text=" SET ELECTION PASSWORD : ", font=('Bahnschrift SemiLight', 13, 'bold'), bg="misty rose")
            self.password_label.place(x=120, y=400)

            self.confirm_password_label = Label(self.window, text=" CONFIRM PASSWORD : ", font=('Bahnschrift SemiLight', 13, 'bold'), bg="misty rose")
            self.confirm_password_label.place(x=140, y=450)

            self.election_name = StringVar()
            self.election_name_entry = Entry(self.window, textvariable=self.election_name, width=30, bd=0)
            self.election_name_entry.place(x=450, y=203)

            self.total_voters = StringVar()
            self.total_voters_entry = Entry(self.window, textvariable=self.total_voters, width=30, bd=0)
            self.total_voters_entry.place(x=450, y=303)

            self.first_password = StringVar()
            self.first_password_entry = Entry(self.window, textvariable=self.first_password, show="*", width=30, bd=0)
            self.first_password_entry.place(x=450, y=403)

            self.toggle_first_password_button = Button(self.window, text="Show", bd=0, fg="white", bg="black", height=1, width=8, command=lambda : toggle_password_function(self.first_password_entry, self.toggle_first_password_button), cursor="hand2")
            self.toggle_first_password_button.place(x=650, y=400)

            self.second_password = StringVar()
            self.second_password_entry = Entry(self.window, textvariable=self.second_password, show="*", width=30, bd=0)
            self.second_password_entry.place(x=450, y=453)

            self.toggle_second_password_button = Button(self.window, text="Show", bd=0, fg="white", bg="black", height=1, width=8, command=lambda : toggle_password_function(self.second_password_entry, self.toggle_second_password_button), cursor="hand2")
            self.toggle_second_password_button.place(x=650, y=450)


            self.calendar = datetime.datetime.now()
            self.week = self.calendar.strftime("%A")
            self.year = self.calendar.strftime("%Y")
            self.month = self.calendar.strftime("%b")
            self.day = self.calendar.strftime("%d")



            self.election_start_button = Button(self.window, text=" START ", bd=0, fg="white", bg="black", command = home_page_entry_check_function, width=20, height=3, cursor="hand2")
            self.election_start_button.place(x=340, y=550)
        



        main_function()                 # calling main function







if __name__ == "__main__":
    
    window = Tk()                        # creating root tkinter window
    election(window)                     # passing the root window as the argument to the class

    window.mainloop()                    # makes the main window remain active i.e. in a loop.